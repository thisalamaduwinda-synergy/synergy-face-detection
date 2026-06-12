"""
attendance_exporter.py
─────────────────────────────────────────────────────────────
Exports daily attendance to CSV + Excel at 11:59 PM (local time).
Optionally emails the report to configured recipients.

Saves to:  reports/attendance_YYYY-MM-DD.csv / .xlsx
"""

from __future__ import annotations

import csv
import io
import logging
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING, Dict, List, Optional

if TYPE_CHECKING:
    from modules.employee_database import EmployeeDatabase

logger = logging.getLogger(__name__)

# ── Palette used for the Excel report ───────────────────────
_XL_HEADER_FILL  = "1E3A5F"   # dark navy header
_XL_HEADER_FONT  = "FFFFFF"   # white text
_XL_ALT_FILL     = "F1F5F9"   # light-blue alternate row
_XL_GREEN        = "16A34A"   # On Time
_XL_RED          = "DC2626"   # Late
_XL_GOLD         = "D97706"   # rate-mid
_XL_RATE_GOOD    = "DCFCE7"   # rate ≥ 80 %
_XL_RATE_MID     = "FEF9C3"   # rate 50-79 %
_XL_RATE_LOW     = "FEE2E2"   # rate < 50 %

# ─────────────────────────────────────────────────────────────
# CSV helpers
# ─────────────────────────────────────────────────────────────

CSV_HEADERS = [
    "Employee ID", "Name", "Department", "Date",
    "First Seen", "Last Seen", "Work Hours", "Status",
    "Camera", "Confidence (%)",
]

MONTHLY_HEADERS = [
    "Employee ID", "Name", "Department",
    "Days Present", "Days Absent", "Late Days",
    "Total Days", "Attendance Rate (%)",
]


def _fmt_duration(minutes: int) -> str:
    h, m = divmod(minutes, 60)
    return f"{h}h {m:02d}m"


def build_csv_bytes(records: List[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_HEADERS, lineterminator="\r\n")
    writer.writeheader()
    for r in records:
        wm = r.get("work_duration_minutes")
        # Slice ISO datetime "2026-05-27T08:30:00" → "08:30:00" (date column has the date)
        first = (r.get("first_seen") or "")[11:19]
        last  = (r.get("last_seen")  or "")[11:19]
        writer.writerow({
            "Employee ID":    r.get("employee_id", ""),
            "Name":           r.get("employee_name", ""),
            "Department":     r.get("department", ""),
            "Date":           r.get("date", ""),
            "First Seen":     first,
            "Last Seen":      last,
            "Work Hours":     _fmt_duration(wm) if wm is not None else "",
            "Status":         "Late" if r.get("is_late") else "On Time",
            "Camera":         r.get("camera_id", ""),
            "Confidence (%)": f"{float(r.get('confidence', 0)) * 100:.1f}",
        })
    return "﻿".encode("utf-8") + buf.getvalue().encode("utf-8")


def build_monthly_csv_bytes(records: List[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=MONTHLY_HEADERS, lineterminator="\r\n")
    writer.writeheader()
    for r in records:
        writer.writerow({
            "Employee ID":        r.get("employee_id", ""),
            "Name":               r.get("name", ""),
            "Department":         r.get("department", ""),
            "Days Present":       r.get("days_present", 0),
            "Days Absent":        r.get("days_absent", 0),
            "Late Days":          r.get("late_days", 0),
            "Total Days":         r.get("total_days", 0),
            "Attendance Rate (%)": f"{r.get('attendance_rate', 0):.1f}",
        })
    return "﻿".encode("utf-8") + buf.getvalue().encode("utf-8")


# ─────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────

def _xl_header_style(ws, headers: List[str]) -> None:
    """Apply bold white-on-navy header to row 1 and freeze it."""
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor=_XL_HEADER_FILL)
    font = Font(bold=True, color=_XL_HEADER_FONT, size=11)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _xl_autowidth(ws) -> None:
    """Set column widths based on content length."""
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def build_excel_bytes(records: List[dict]) -> bytes:
    """Return .xlsx bytes for a daily attendance report (formatted)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.row_dimensions[1].height = 22

    _xl_header_style(ws, CSV_HEADERS)

    alt_fill   = PatternFill("solid", fgColor=_XL_ALT_FILL)
    green_font = Font(bold=True, color=_XL_GREEN)
    red_font   = Font(bold=True, color=_XL_RED)

    for row_idx, r in enumerate(records, start=2):
        wm     = r.get("work_duration_minutes")
        is_late = r.get("is_late", False)
        row_data = [
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            r.get("department", ""),
            r.get("date", ""),
            (r.get("first_seen") or "")[:19].replace("T", " "),
            (r.get("last_seen")  or "")[:19].replace("T", " "),
            _fmt_duration(wm) if wm is not None else "",
            "Late" if is_late else "On Time",
            r.get("camera_id", ""),
            round(float(r.get("confidence", 0)) * 100, 1),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            # Status column (col 8) colour
            if col_idx == 8:
                cell.font = red_font if is_late else green_font
                cell.alignment = Alignment(horizontal="center")

    _xl_autowidth(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_monthly_excel_bytes(records: List[dict]) -> bytes:
    """Return .xlsx bytes for a monthly attendance summary (formatted)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Summary"
    ws.row_dimensions[1].height = 22

    _xl_header_style(ws, MONTHLY_HEADERS)

    alt_fill = PatternFill("solid", fgColor=_XL_ALT_FILL)
    fills = {
        "good": PatternFill("solid", fgColor=_XL_RATE_GOOD),
        "mid":  PatternFill("solid", fgColor=_XL_RATE_MID),
        "low":  PatternFill("solid", fgColor=_XL_RATE_LOW),
    }
    bold = Font(bold=True)

    for row_idx, r in enumerate(records, start=2):
        rate = r.get("attendance_rate", 0)
        row_data = [
            r.get("employee_id", ""),
            r.get("name", ""),
            r.get("department", ""),
            r.get("days_present", 0),
            r.get("days_absent", 0),
            r.get("late_days", 0),
            r.get("total_days", 0),
            rate,
        ]
        rate_fill = fills["good"] if rate >= 80 else fills["mid"] if rate >= 50 else fills["low"]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0 and col_idx != 8:
                cell.fill = alt_fill
            if col_idx == 8:
                cell.fill = rate_fill
                cell.font = bold
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "0.0%"
                cell.value = rate / 100   # store as fraction for % format

    _xl_autowidth(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# Scheduler
# ─────────────────────────────────────────────────────────────

def _seconds_until(hour: int, minute: int, second: int = 0) -> float:
    now = datetime.now()
    target = now.replace(hour=hour, minute=minute, second=second, microsecond=0)
    if target <= now:
        target += timedelta(days=1)
    return (target - now).total_seconds()


class AttendanceExporter:
    """
    Background thread that saves CSV + Excel at 11:59 PM and
    optionally emails the report.
    """

    def __init__(
        self,
        db: "EmployeeDatabase",
        reports_dir: str = "reports",
        export_hour: int = 23,
        export_minute: int = 59,
        email_service=None,   # EmailService instance or None
    ) -> None:
        self._db           = db
        self._reports_dir  = Path(reports_dir)
        self._export_hour  = export_hour
        self._export_minute = export_minute
        self._email_svc    = email_service
        self._running      = False
        self._thread: Optional[threading.Thread] = None

    def start(self) -> None:
        self._reports_dir.mkdir(parents=True, exist_ok=True)
        self._running = True
        self._thread = threading.Thread(
            target=self._scheduler_loop,
            daemon=True,
            name="attendance-exporter",
        )
        self._thread.start()
        logger.info(
            "Attendance exporter started — daily export at %02d:%02d",
            self._export_hour, self._export_minute,
        )

    def stop(self) -> None:
        self._running = False

    def export_now(self, target_date=None) -> Path:
        """Export CSV + Excel for *target_date*. Returns the CSV path."""
        from datetime import date as _date
        target_date = target_date or _date.today()
        records = self._db.get_attendance_by_date(target_date)

        csv_bytes = build_csv_bytes(records)
        csv_path  = self._reports_dir / f"attendance_{target_date}.csv"
        csv_path.write_bytes(csv_bytes)

        try:
            xl_bytes = build_excel_bytes(records)
            xl_path  = self._reports_dir / f"attendance_{target_date}.xlsx"
            xl_path.write_bytes(xl_bytes)
            logger.info("Attendance Excel saved: %s", xl_path)
        except Exception:
            logger.exception("Excel export failed — CSV still saved")
            xl_bytes = None

        logger.info("Attendance CSV saved: %s (%d records)", csv_path, len(records))
        return csv_path

    def _scheduler_loop(self) -> None:
        while self._running:
            wait = _seconds_until(self._export_hour, self._export_minute)
            logger.debug("Next attendance export in %.0f seconds", wait)

            elapsed = 0.0
            while self._running and elapsed < wait:
                chunk = min(30.0, wait - elapsed)
                time.sleep(chunk)
                elapsed += chunk

            if not self._running:
                break

            try:
                self._run_daily_export()
            except Exception:
                logger.exception("Daily attendance export failed")

            time.sleep(70)   # skip past midnight before recalculating

    def _run_daily_export(self) -> None:
        from datetime import date as _date
        today    = _date.today()
        records  = self._db.get_attendance_by_date(today)
        csv_bytes = build_csv_bytes(records)

        # Save CSV to reports/
        csv_path = self._reports_dir / f"attendance_{today}.csv"
        csv_path.write_bytes(csv_bytes)

        # Save Excel to reports/
        xl_bytes: Optional[bytes] = None
        xl_name = f"attendance_{today}.xlsx"
        try:
            xl_bytes = build_excel_bytes(records)
            (self._reports_dir / xl_name).write_bytes(xl_bytes)
        except Exception:
            logger.exception("Excel export failed during daily run")

        # Also copy both files to the Desktop for easy access
        desktop = Path.home() / "Desktop"
        try:
            desktop_csv = desktop / f"attendance_{today}.csv"
            desktop_csv.write_bytes(csv_bytes)
            logger.info("Attendance CSV saved to Desktop: %s", desktop_csv)
        except Exception:
            logger.warning("Could not save CSV to Desktop")

        if xl_bytes is not None:
            try:
                desktop_xl = desktop / xl_name
                desktop_xl.write_bytes(xl_bytes)
                logger.info("Attendance Excel saved to Desktop: %s", desktop_xl)
            except Exception:
                logger.warning("Could not save Excel to Desktop")

        logger.info("Daily export done: %d records for %s", len(records), today)

        # Send email
        if self._email_svc and self._email_svc.enabled:
            absent = self._db.get_absent_employees(today)
            self._email_svc.send_daily_report(
                target_date=today,
                records=records,
                absent=absent,
                csv_bytes=csv_bytes,
                excel_bytes=xl_bytes,
            )
